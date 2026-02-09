# #file handling in python

# file = open("sample.txt", "w")
# file.write("Hello, this is a file handling example.")
# file.close()

# file = open("sample.txt", "r")
# content = file.read()
# print(content)
# file.close()

# with open("sample.txt", "r") as file:
#     content = file.read()
#     print(content)


# # Writing to a file With file not found error handling

# try:
#     with open("missing.txt", "r") as file:
#         print(file.read())
# except FileNotFoundError:
#     print("File not found. Please check the filename.")


# # reading a file line by line by data.csv file 

# import csv
# with open("data.csv","r") as file:
#     reader = csv.reader(file)
#     for row in reader :
#         print(row[0], row[1], row[2])


# importing file using Excel 

from openpyxl import load_workbook

wb = load_workbook("info.xlsx")
sheet = wb.active

for row in sheet.iter_rows(values_only=True):
    print(row[0], row[1], row[2])

    





